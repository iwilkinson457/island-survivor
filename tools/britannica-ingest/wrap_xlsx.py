"""
XLSX / XLSM / XLS -> Markdown wrapper for Britannica ingestion.

Uses openpyxl (already installed). XLS files require xlrd (not installed by default)
and will be reported as unsupported if xlrd is absent.

Extracts:
- Sheet names
- Up to 3 sheets, first 50 rows each (configurable)
- Column headers as markdown table headers
- Named ranges (if any)

Limitations:
- Formulas are not evaluated — cell values shown are the last calculated result.
- Charts and images are ignored.
- Very large sheets are truncated.
"""

from pathlib import Path
import re

MAX_SHEETS = 3
MAX_ROWS = 50
MAX_COLS = 12


def extract(source: Path) -> dict:
    ext = source.suffix.lower()

    if ext == ".xls":
        return _extract_xls_stub(source)

    try:
        import openpyxl
    except ImportError:
        return {"error": "openpyxl not installed (pip install openpyxl)"}

    try:
        wb = openpyxl.load_workbook(str(source), read_only=True, data_only=True)
    except Exception as e:
        return {"error": f"openpyxl failed to open workbook: {e}"}

    sheet_names = wb.sheetnames
    body_parts = []
    all_cell_text = []
    tables_extracted = []

    skipped_chart_sheets = []

    for sheet_name in sheet_names[:MAX_SHEETS]:
        ws = wb[sheet_name]

        # Chartsheets have no rows/cells — skip gracefully
        from openpyxl.chartsheet.chartsheet import Chartsheet
        if isinstance(ws, Chartsheet):
            skipped_chart_sheets.append(sheet_name)
            continue

        rows_data = []

        for i, row in enumerate(ws.iter_rows(max_row=MAX_ROWS, max_col=MAX_COLS, values_only=True)):
            cells = [_cell_str(c) for c in row]
            if any(c for c in cells):  # skip fully empty rows
                rows_data.append(cells)
            if i >= MAX_ROWS:
                break

        if not rows_data:
            body_parts.append(f"### Sheet: {sheet_name}\n\n_(empty)_")
            continue

        # Normalise column count
        max_cols = max(len(r) for r in rows_data)
        rows_data = [(r + [""] * max_cols)[:max_cols] for r in rows_data]

        header = rows_data[0]
        md_rows = [
            f"| {' | '.join(header)} |",
            f"| {' | '.join(['---'] * max_cols)} |",
        ]
        for row in rows_data[1:]:
            md_rows.append(f"| {' | '.join(row)} |")

        table_md = "\n".join(md_rows)
        body_parts.append(f"### Sheet: {sheet_name}\n\n{table_md}")

        for row in rows_data:
            all_cell_text.extend([c for c in row if c])

        tables_extracted.append({
            "sheet": sheet_name,
            "rows": len(rows_data) - 1,
            "cols": max_cols,
            "header": header[:6],
        })

    wb.close()

    # Title: workbook filename
    title = source.stem

    # Summary
    summary = _build_summary(sheet_names, tables_extracted, source)

    # Key facts
    key_facts = _extract_key_facts(tables_extracted, sheet_names, all_cell_text)

    # Operational relevance
    op_relevance = _detect_operational_relevance(" ".join(all_cell_text))

    # Gaps
    gaps = []
    if len(sheet_names) > MAX_SHEETS:
        gaps.append(f"Only first {MAX_SHEETS} of {len(sheet_names)} sheets extracted")
    for te in tables_extracted:
        if te["rows"] >= MAX_ROWS - 1:
            gaps.append(f"Sheet '{te['sheet']}' truncated at {MAX_ROWS} rows -- may have more data")
    if skipped_chart_sheets:
        gaps.append(f"Chart sheet(s) skipped (no row data): {', '.join(skipped_chart_sheets)}")
    if not tables_extracted:
        if skipped_chart_sheets and len(skipped_chart_sheets) == len(sheet_names[:MAX_SHEETS]):
            gaps.append("All examined sheets were chart sheets -- no cell data extracted")
        else:
            gaps.append("No usable cell data found in examined sheets (sheets may be empty or chart-only)")

    body_md = "\n\n".join(body_parts)

    return {
        "title": title,
        "summary": summary,
        "key_facts": key_facts,
        "operational_relevance": op_relevance,
        "gaps": gaps,
        "method": f"openpyxl (max {MAX_SHEETS} sheets, {MAX_ROWS} rows each)",
        "body_md": body_md,
    }


def _extract_xls_stub(source: Path) -> dict:
    """Attempt XLS extraction via xlrd if available, otherwise stub."""
    try:
        import xlrd
        wb = xlrd.open_workbook(str(source))
        sheet_names = wb.sheet_names()
        rows_extracted = []
        for sname in sheet_names[:MAX_SHEETS]:
            ws = wb.sheet_by_name(sname)
            for rx in range(min(MAX_ROWS, ws.nrows)):
                row = [str(ws.cell_value(rx, cx)) for cx in range(min(MAX_COLS, ws.ncols))]
                rows_extracted.extend([c for c in row if c.strip()])
        summary = f"Legacy XLS workbook with sheets: {', '.join(sheet_names)}"
        key_facts = [f"Sheet '{s}'" for s in sheet_names[:6]]
        key_facts += rows_extracted[:6]
        return {
            "title": source.stem,
            "summary": summary,
            "key_facts": key_facts,
            "operational_relevance": "Legacy XLS format — review content manually",
            "gaps": ["XLS extracted via xlrd — formula results and formatting not preserved"],
            "method": "xlrd",
        }
    except ImportError:
        return {
            "title": source.stem,
            "summary": "",
            "key_facts": [],
            "operational_relevance": "(not determined)",
            "gaps": [
                "XLS format is not natively supported by openpyxl.",
                "Install xlrd (pip install xlrd) for legacy XLS support, "
                "or convert to XLSX in Excel first.",
            ],
            "error": "XLS requires xlrd — not installed",
        }
    except Exception as e:
        return {"error": f"xlrd extraction failed: {e}"}


def _cell_str(value) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    # Limit cell width for readability
    return s[:80] if len(s) > 80 else s


def _build_summary(sheet_names: list, tables: list, source: Path) -> str:
    if not tables:
        return f"Workbook '{source.name}' contains {len(sheet_names)} sheet(s) but no readable data was extracted."
    parts = [f"Workbook '{source.name}' with {len(sheet_names)} sheet(s)."]
    for t in tables:
        cols_str = ", ".join(str(c) for c in t["header"] if c)
        parts.append(
            f"Sheet '{t['sheet']}': {t['rows']} data rows, {t['cols']} columns"
            + (f" — headers: {cols_str}" if cols_str else "")
        )
    return " ".join(parts)


def _extract_key_facts(tables: list, sheet_names: list, cell_text: list) -> list:
    facts = []
    facts.append(f"Sheets: {', '.join(sheet_names[:8])}")
    for t in tables:
        header_str = ", ".join(str(c) for c in t["header"] if c)
        if header_str:
            facts.append(f"'{t['sheet']}' columns include: {header_str}")
        facts.append(f"'{t['sheet']}' has {t['rows']} data rows")
    # Add a few sample values
    samples = [c for c in cell_text if len(c) > 3 and len(c) < 60][:4]
    for s in samples:
        facts.append(f"Sample value: {s}")
    return facts[:10]


def _detect_operational_relevance(body: str) -> str:
    lower = body.lower()
    signals = {
        "parameter": "Contains parameter data",
        "setpoint": "Contains setpoint/configuration values",
        "alarm": "References alarm data",
        "maintenance": "Maintenance data",
        "schedule": "Schedule or planning data",
        "budget": "Budget or financial data",
        "specification": "Specification data",
        "test": "Test or verification data",
        "calibration": "Calibration data",
    }
    matched = [v for k, v in signals.items() if k in lower]
    return "; ".join(matched[:3]) if matched else "Spreadsheet data — review sheets for specific relevance"
